import io
import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & STYLES
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Cashfree × Shopify ", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #F8FAFC; }
    .main-header {
        background: #1F3864; padding: 25px; border-radius: 12px; 
        color: white; margin-bottom: 25px; text-align: center;
    }
    .metric-card {
        background: white; border-radius: 10px; padding: 15px;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;
        border-bottom: 4px solid #1F3864;
    }
    .metric-val { font-size: 24px; font-weight: 800; color: #1F3864; }
    .metric-lbl { font-size: 11px; color: #64748B; text-transform: uppercase; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SMART HEADER DETECTION
# ─────────────────────────────────────────────────────────────────────────────
def smart_read_excel(file_bytes, anchor_cols):
    try:
        preview = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=100)
        header_row_index = None
        anchors = [str(a).strip().lower() for a in anchor_cols]

        for i, row in preview.iterrows():
            row_values = [str(val).strip().lower() for val in row.values if pd.notna(val)]
            matches = sum(1 for a in anchors if any(a in val for val in row_values))
            if matches >= 2: 
                header_row_index = i
                break
        
        if header_row_index is None: return None

        df = pd.read_excel(io.BytesIO(file_bytes), header=header_row_index)
        df.columns = [str(c).strip() for c in df.columns]
        
        rename_map = {}
        for actual_col in df.columns:
            for target in anchor_cols:
                if target.lower() in actual_col.lower():
                    rename_map[actual_col] = target
        return df.rename(columns=rename_map)
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def get_excel_styles():
    thin = Side(style='thin', color='AAAAAA')
    return {
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'al_c': Alignment(horizontal='center', vertical='center'),
        'hdr_fill': PatternFill('solid', fgColor='1F3864'),
        'hdr_font': Font(name='Arial', bold=True, color='FFFFFF', size=11),
        'credit_fill': PatternFill('solid', fgColor='E2EFDA'),
        'debit_fill': PatternFill('solid', fgColor='FCE4D6')
    }

def build_journal_xlsx(merged):
    wb = Workbook()
    ws = wb.active
    styles = get_excel_styles()
    sorted_df = merged.sort_values(by='_sort_priority')
    
    headers = ['Settlement Date', 'Credit Account', 'Debit Account', 'Order Number', 'Amount', 'Reference ID']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = styles['hdr_fill'], styles['hdr_font'], styles['al_c'], styles['border']
        ws.column_dimensions[get_column_letter(ci)].width = 25

    for ri, row in enumerate(sorted_df.to_dict('records'), 2):
        is_credit = row.get('Sale Type') == 'CREDIT'
        row_fill = styles['credit_fill'] if is_credit else styles['debit_fill']
        
        date_cell = ws.cell(row=ri, column=1, value=row.get('Settlement Date'))
        date_cell.number_format = 'DD.MM.YYYY'
        
        cells = [
            date_cell,
            ws.cell(row=ri, column=2, value=row.get('Customer Email') if is_credit else 'Cashfree Receivable'),
            ws.cell(row=ri, column=3, value='Cashfree Receivable' if is_credit else row.get('Customer Email')),
            ws.cell(row=ri, column=4, value=str(row.get('Order Number', 'N/A'))),
            ws.cell(row=ri, column=5, value=row.get('Event Amount')),
            ws.cell(row=ri, column=6, value=row.get('Merchant Reference Id'))
        ]
        for cell in cells:
            cell.border, cell.fill = styles['border'], row_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# MAIN INTERFACE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header"><h1>📊 Cashfree × Shopify Reconciliation</h1></div>', unsafe_allow_html=True)

# File Upload Section
u1, u2 = st.columns(2)
with u1:
    cf_file = st.file_uploader("Upload Cashfree Settlement Report", type=['xlsx'], key="cf")
with u2:
    sh_file = st.file_uploader("Upload Shopify Order Export", type=['xlsx'], key="sh")

# New Filename Input
custom_filename = st.text_input("Output Journal File Name", value=f"Journal_{datetime.date.today()}")

# Session State Initialization to keep data persistent
if 'merged_result' not in st.session_state:
    st.session_state.merged_result = None

if cf_file and sh_file:
    st.divider()
    
    if st.button("▶ Run Reconciliation", type="primary", use_container_width=True):
        with st.spinner("Analyzing headers and merging transactions..."):
            cf_anchors = ['Event Type', 'Sale Type', 'Customer Email', 'Event Amount', 'Merchant Reference Id', 'Settlement Date']
            sh_anchors = ['Order Number', 'Email']
            
            cf_df = smart_read_excel(cf_file.getvalue(), cf_anchors)
            sh_df = smart_read_excel(sh_file.getvalue(), sh_anchors)

            if cf_df is not None and sh_df is not None:
                # Clean Cashfree
                cf_df = cf_df[cf_df['Event Type'].astype(str).str.upper().isin(['PAYMENT', 'REFUND'])].copy()
                cf_df['_email_key'] = cf_df['Customer Email'].fillna('').astype(str).str.strip().str.lower()
                cf_df['_sort_priority'] = cf_df['Sale Type'].apply(lambda x: 0 if str(x).upper() == 'CREDIT' else 1)
                cf_df['Settlement Date'] = pd.to_datetime(cf_df['Settlement Date'], errors='coerce').dt.date
                
                # Clean Shopify
                sh_df['_email_key'] = sh_df['Email'].fillna('').astype(str).str.strip().str.lower()
                sh_subset = sh_df[['_email_key', 'Order Number']].drop_duplicates(subset=['_email_key'])

                # Save to session state
                st.session_state.merged_result = cf_df.merge(sh_subset, on='_email_key', how='left')
                st.success("Reconciliation Complete.")
            else:
                st.error("Header detection failed. Please check the file structure.")

    # Show results and download button if data exists in memory
    if st.session_state.merged_result is not None:
        merged = st.session_state.merged_result
        
        m1, m2, m3, m4 = st.columns(4)
        m1.markdown(f'<div class="metric-card"><div class="metric-val">{len(merged)}</div><div class="metric-lbl">Total</div></div>', unsafe_allow_html=True)
        m2.markdown(f'<div class="metric-card"><div class="metric-val">{len(merged[merged["_sort_priority"]==0])}</div><div class="metric-lbl">Payments</div></div>', unsafe_allow_html=True)
        m3.markdown(f'<div class="metric-card"><div class="metric-val">{len(merged[merged["_sort_priority"]==1])}</div><div class="metric-lbl">Refunds</div></div>', unsafe_allow_html=True)
        m4.markdown(f'<div class="metric-card"><div class="metric-val" style="color:#E11D48">{merged["Order Number"].isna().sum()}</div><div class="metric-lbl">Unmatched Email</div></div>', unsafe_allow_html=True)

        # Ensure correct extension
        fn = custom_filename if custom_filename.endswith(".xlsx") else f"{custom_filename}.xlsx"

        st.download_button(
            label="⬇️ Download Journal with Settlement Dates", 
            data=build_journal_xlsx(merged), 
            file_name=fn, 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )
else:
    # Clear session if files are removed
    st.session_state.merged_result = None
    st.info("Upload your reports to begin.")
